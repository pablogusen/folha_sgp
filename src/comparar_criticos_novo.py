import json
import pandas as pd
import re
from unidecode import unidecode
from collections import Counter
from pathlib import Path

def normalizar_nome(nome):
    """Normaliza nome removendo acentos, pontos, hífens e espaços extras"""
    if not nome or pd.isna(nome):
        return ""
    # Remove acentos
    nome = unidecode(str(nome))
    # Remove pontos, hífens e caracteres especiais
    nome = re.sub(r'[.\-/,]', ' ', nome)
    # Remove espaços extras e converte para maiúsculas
    nome = ' '.join(nome.upper().split())
    return nome

print("="*80)
print("🔍 COMPARAÇÃO: RELATÓRIO CONSIGNO vs SERVIDORES CRÍTICOS")
print("="*80)

# 1. Carregar dados do JSON
caminho_base = Path(__file__).parent.parent
caminho_json = caminho_base / 'data' / 'backup' / 'dados_folhas_backup.json'

with open(caminho_json, 'r', encoding='utf-8') as f:
    dados_folhas = json.load(f)

# Filtrar os críticos usando o MESMO CRITÉRIO do relatório HTML (181 pessoas)
# Critério: Descontos Extras > 35% do LÍQUIDO FINAL (não da margem consignável)
servidores_criticos = []
for servidor in dados_folhas:
    descontos_extras = servidor.get('total_descontos_extras', 0)
    liquido_final = servidor.get('liquido', 0)
    
    if liquido_final > 0:
        percentual = (descontos_extras / liquido_final) * 100
        
        if percentual > 35:
            # Calcular também a margem consignável para referência
            margem_consignavel = servidor.get('total_proventos', 0) - servidor.get('total_descontos_obrigatorios', 0)
            
            servidores_criticos.append({
                'Nome_Original': servidor.get('nome', ''),
                'Nome_Normalizado': normalizar_nome(servidor.get('nome', '')),
                'CPF': servidor.get('cpf', ''),
                'Matrícula': servidor.get('matricula', ''),
                'Situação': servidor.get('situacao', 'N/A'),
                'Percentual': round(percentual, 1),
                'Liquido_Final': round(liquido_final, 2),
                'Margem_Consignavel': round(margem_consignavel, 2),
                'Descontos_Extras': round(descontos_extras, 2)
            })

print(f"\n📊 Total de servidores CRÍTICOS no ESTUDO (Descontos > 35% do Líquido): {len(servidores_criticos)}")

# Contar por situação
situacoes = Counter([s['Situação'] for s in servidores_criticos])
print(f"\n📋 Distribuição por situação:")
for situacao, qtd in situacoes.items():
    print(f"   {situacao}: {qtd}")

# 2. Carregar planilha do relatório (aba "Margens")
df_relatorio = pd.read_excel('Relatorio-Consigno-MargemConsignavel-07122025.xlsx', sheet_name='Margens')

print(f"\n📄 Total de registros na PLANILHA EXTERNA: {len(df_relatorio)}")
print(f"📊 TOTAL ESPERADO: 181 (estudo) + {len(df_relatorio)} (planilha) = {181 + len(df_relatorio)} nomes")

# 3. Criar dicionários com nomes normalizados
criticos_dict = {}
for s in servidores_criticos:
    nome_norm = s['Nome_Normalizado']
    if nome_norm:  # Evitar nomes vazios
        criticos_dict[nome_norm] = s

relatorio_dict = {}
for idx, row in df_relatorio.iterrows():
    nome_norm = normalizar_nome(row.get('Funcionário', ''))  # Coluna com acento
    if nome_norm:  # Evitar nomes vazios
        relatorio_dict[nome_norm] = {
            'Nome_Original': row.get('Funcionário', ''),
            'Matrícula': row.get('Matrícula', ''),
            'Observação': row.get('Observação', '')
        }

print(f"\n🔍 Nomes únicos nos CRÍTICOS: {len(criticos_dict)}")
print(f"🔍 Nomes únicos na PLANILHA: {len(relatorio_dict)}")

# 4. COMPARAÇÃO: Identificar AMBOS, APENAS ESTUDO, APENAS PLANILHA
print("\n" + "="*80)
print("📋 DISTRIBUIÇÃO DOS 281 NOMES")
print("="*80)

# Identificar quais nomes estão em AMBOS (interseção)
nomes_em_ambos = set(criticos_dict.keys()) & set(relatorio_dict.keys())

print(f"\n🔍 Nomes em AMBAS as listas: {len(nomes_em_ambos)}")

# ABA 1: AMBOS - Nomes que aparecem nas DUAS fontes (mostrar dados do estudo)
ambos = []
for nome_norm in nomes_em_ambos:
    dados_estudo = criticos_dict[nome_norm]
    dados_planilha = relatorio_dict[nome_norm]
    ambos.append({
        'Nome': dados_estudo['Nome_Original'],
        'Matrícula': dados_estudo['Matrícula'],
        'CPF': dados_estudo['CPF'],
        'Situação': dados_estudo['Situação'],
        'Percentual_Critico': dados_estudo['Percentual'],
        'Liquido_Final': dados_estudo['Liquido_Final'],
        'Margem_Consignavel': dados_estudo['Margem_Consignavel'],
        'Descontos_Extras': dados_estudo['Descontos_Extras'],
        'Observação_Planilha': dados_planilha['Observação']
    })

# ABA 2: APENAS ESTUDO - Todos os 181 do estudo que NÃO estão em AMBOS
apenas_estudo = []
for nome_norm, dados in criticos_dict.items():
    if nome_norm not in nomes_em_ambos:
        apenas_estudo.append({
            'Nome': dados['Nome_Original'],
            'Matrícula': dados['Matrícula'],
            'CPF': dados['CPF'],
            'Situação': dados['Situação'],
            'Percentual_Critico': dados['Percentual'],
            'Liquido_Final': dados['Liquido_Final'],
            'Margem_Consignavel': dados['Margem_Consignavel'],
            'Descontos_Extras': dados['Descontos_Extras']
        })

# ABA 3: APENAS PLANILHA - Todos da planilha que NÃO estão em AMBOS
apenas_planilha = []
for nome_norm, dados in relatorio_dict.items():
    if nome_norm not in nomes_em_ambos:
        apenas_planilha.append({
            'Nome': dados['Nome_Original'],
            'Matrícula': dados['Matrícula'],
            'Observação': dados['Observação']
        })

total_distribuido = len(ambos) + len(apenas_estudo) + len(apenas_planilha)
total_esperado = len(criticos_dict) + len(relatorio_dict)

print(f"\n✅ AMBOS: {len(ambos)} nomes")
print(f"🔴 APENAS ESTUDO: {len(apenas_estudo)} nomes (181 - {len(ambos)} = {len(apenas_estudo)})")
print(f"🔵 APENAS PLANILHA: {len(apenas_planilha)} nomes ({len(relatorio_dict)} - {len(ambos)} = {len(apenas_planilha)})")
print(f"\n📊 TOTAL DISTRIBUÍDO: {total_distribuido}")
print(f"📊 TOTAL ESPERADO: 181 + {len(relatorio_dict)} = {total_esperado}")
print(f"\n✔️  Verificação: {len(ambos)} + {len(apenas_estudo)} + {len(apenas_planilha)} = {total_distribuido}")

# 5. Gerar planilha Excel com os resultados - FORMATADA PARA APRESENTAÇÃO
output_file = 'Comparacao_Criticos_vs_Relatorio.xlsx'

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Aba 1: Resumo
    resumo_data = [
        {'Categoria': 'COMPARAÇÃO: ESTUDO (181) vs PLANILHA EXTERNA', 'Quantidade': ''},
        {'Categoria': '', 'Quantidade': ''},
        {'Categoria': '📊 TOTAIS', 'Quantidade': ''},
        {'Categoria': 'Servidores CRÍTICOS no ESTUDO', 'Quantidade': len(servidores_criticos)},
        {'Categoria': 'Servidores na PLANILHA EXTERNA', 'Quantidade': len(df_relatorio)},
        {'Categoria': 'TOTAL (Estudo + Planilha)', 'Quantidade': len(servidores_criticos) + len(df_relatorio)},
        {'Categoria': '', 'Quantidade': ''},
        {'Categoria': '🎯 DISTRIBUIÇÃO (181 + Planilha)', 'Quantidade': ''},
        {'Categoria': '✅ AMBOS (nas duas listas)', 'Quantidade': len(ambos)},
        {'Categoria': '🔴 APENAS ESTUDO (181 - AMBOS)', 'Quantidade': len(apenas_estudo)},
        {'Categoria': '🔵 APENAS PLANILHA (Planilha - AMBOS)', 'Quantidade': len(apenas_planilha)},
        {'Categoria': '', 'Quantidade': ''},
        {'Categoria': '✔️  VERIFICAÇÃO DA SOMA', 'Quantidade': ''},
        {'Categoria': f'   AMBOS + APENAS ESTUDO + APENAS PLANILHA', 'Quantidade': len(ambos) + len(apenas_estudo) + len(apenas_planilha)},
        {'Categoria': f'   Esperado (181 + {len(df_relatorio)})', 'Quantidade': len(servidores_criticos) + len(df_relatorio)},
    ]
    
    df_resumo = pd.DataFrame(resumo_data)
    df_resumo.to_excel(writer, sheet_name='📊 Resumo', index=False, startrow=0)
    
    ws_resumo = writer.sheets['📊 Resumo']
    
    # Formatação do Resumo
    ws_resumo.column_dimensions['A'].width = 55
    ws_resumo.column_dimensions['B'].width = 18
    
    # Título principal
    ws_resumo['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws_resumo['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws_resumo['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Cabeçalhos de seção
    for row in [3, 8, 13]:
        ws_resumo[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
        ws_resumo[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Números em destaque
    for row in [4, 5, 6, 9, 10, 11, 14, 15]:
        ws_resumo[f'B{row}'].font = Font(size=11, bold=True)
        ws_resumo[f'B{row}'].alignment = Alignment(horizontal='center')
    
    # Aba 2: AMBOS (formatada)
    if ambos:
        df_ambos_formatado = pd.DataFrame({
            'Nome do Servidor': [s['Nome'] for s in ambos],
            'Matrícula': [s['Matrícula'] for s in ambos],
            'CPF': [s['CPF'] for s in ambos],
            'Situação': [s['Situação'] for s in ambos],
            '% Crítico': [s['Percentual_Critico'] for s in ambos],
            'Líquido Final (R$)': [s['Liquido_Final'] for s in ambos],
            'Descontos Extras (R$)': [s['Descontos_Extras'] for s in ambos],
            'Margem Consignável (R$)': [s['Margem_Consignavel'] for s in ambos],
            'Status na Planilha': [s['Observação_Planilha'] for s in ambos]
        })
        df_ambos_formatado.to_excel(writer, sheet_name='✅ AMBOS', index=False)
        
        ws_ambos = writer.sheets['✅ AMBOS']
        
        # Formatação de cabeçalho
        for col in range(1, 10):
            cell = ws_ambos.cell(1, col)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar larguras
        ws_ambos.column_dimensions['A'].width = 40
        ws_ambos.column_dimensions['B'].width = 12
        ws_ambos.column_dimensions['C'].width = 15
        ws_ambos.column_dimensions['D'].width = 12
        ws_ambos.column_dimensions['E'].width = 12
        ws_ambos.column_dimensions['F'].width = 18
        ws_ambos.column_dimensions['G'].width = 20
        ws_ambos.column_dimensions['H'].width = 22
        ws_ambos.column_dimensions['I'].width = 20
        
        # Formatar percentuais e valores
        for row in range(2, len(df_ambos_formatado) + 2):
            ws_ambos[f'E{row}'].number_format = '0.0"%"'
            ws_ambos[f'E{row}'].alignment = Alignment(horizontal='center')
            ws_ambos[f'F{row}'].number_format = 'R$ #,##0.00'
            ws_ambos[f'G{row}'].number_format = 'R$ #,##0.00'
            ws_ambos[f'H{row}'].number_format = 'R$ #,##0.00'
            ws_ambos[f'I{row}'].alignment = Alignment(horizontal='center')
            
            # Destaque vermelho para % > 50%
            if ws_ambos[f'E{row}'].value and float(str(ws_ambos[f'E{row}'].value).replace('%','')) > 50:
                ws_ambos[f'E{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                ws_ambos[f'E{row}'].font = Font(bold=True, color='9C0006')
    else:
        pd.DataFrame([{'Mensagem': 'Nenhuma correspondência encontrada'}]).to_excel(
            writer, sheet_name='✅ AMBOS', index=False
        )
    
    # Aba 3: APENAS ESTUDO (formatada)
    if apenas_estudo:
        df_estudo_formatado = pd.DataFrame({
            'Nome do Servidor': [s['Nome'] for s in apenas_estudo],
            'Matrícula': [s['Matrícula'] for s in apenas_estudo],
            'CPF': [s['CPF'] for s in apenas_estudo],
            'Situação': [s['Situação'] for s in apenas_estudo],
            '% Crítico': [s['Percentual_Critico'] for s in apenas_estudo],
            'Líquido Final (R$)': [s['Liquido_Final'] for s in apenas_estudo],
            'Descontos Extras (R$)': [s['Descontos_Extras'] for s in apenas_estudo],
            'Margem Consignável (R$)': [s['Margem_Consignavel'] for s in apenas_estudo]
        })
        df_estudo_formatado.to_excel(writer, sheet_name='🔴 APENAS ESTUDO', index=False)
        
        ws_estudo = writer.sheets['🔴 APENAS ESTUDO']
        
        # Formatação de cabeçalho
        for col in range(1, 9):
            cell = ws_estudo.cell(1, col)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar larguras
        ws_estudo.column_dimensions['A'].width = 40
        ws_estudo.column_dimensions['B'].width = 12
        ws_estudo.column_dimensions['C'].width = 15
        ws_estudo.column_dimensions['D'].width = 12
        ws_estudo.column_dimensions['E'].width = 12
        ws_estudo.column_dimensions['F'].width = 18
        ws_estudo.column_dimensions['G'].width = 20
        ws_estudo.column_dimensions['H'].width = 22
        
        # Formatar percentuais e valores
        for row in range(2, len(df_estudo_formatado) + 2):
            ws_estudo[f'E{row}'].number_format = '0.0"%"'
            ws_estudo[f'E{row}'].alignment = Alignment(horizontal='center')
            ws_estudo[f'F{row}'].number_format = 'R$ #,##0.00'
            ws_estudo[f'G{row}'].number_format = 'R$ #,##0.00'
            ws_estudo[f'H{row}'].number_format = 'R$ #,##0.00'
            
            # Destaque laranja para % > 60%
            if ws_estudo[f'E{row}'].value and float(str(ws_estudo[f'E{row}'].value).replace('%','')) > 60:
                ws_estudo[f'E{row}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                ws_estudo[f'E{row}'].font = Font(bold=True)
    else:
        pd.DataFrame([{'Mensagem': 'Todos os críticos estão na planilha'}]).to_excel(
            writer, sheet_name='🔴 APENAS ESTUDO', index=False
        )
    
    # Aba 4: APENAS PLANILHA (formatada)
    if apenas_planilha:
        df_planilha_formatado = pd.DataFrame({
            'Nome do Servidor': [s['Nome'] for s in apenas_planilha],
            'Matrícula': [s['Matrícula'] for s in apenas_planilha],
            'Observação': [s['Observação'] for s in apenas_planilha]
        })
        df_planilha_formatado.to_excel(writer, sheet_name='🔵 APENAS PLANILHA', index=False)
        
        ws_planilha = writer.sheets['🔵 APENAS PLANILHA']
        
        # Formatação de cabeçalho
        for col in range(1, 4):
            cell = ws_planilha.cell(1, col)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='2E75B5', end_color='2E75B5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar larguras
        ws_planilha.column_dimensions['A'].width = 40
        ws_planilha.column_dimensions['B'].width = 12
        ws_planilha.column_dimensions['C'].width = 20
    else:
        pd.DataFrame([{'Mensagem': 'Todos da planilha são críticos'}]).to_excel(
            writer, sheet_name='🔵 APENAS PLANILHA', index=False
        )

print(f"\n✅ Planilha gerada: {output_file}")

print("\n" + "="*80)
print("📋 EXEMPLOS (Primeiros 5 de cada categoria)")
print("="*80)

if ambos:
    print("\n✅ AMBOS (Estudo + Planilha):")
    for i, item in enumerate(ambos[:5], 1):
        print(f"{i}. {item['Nome']}")
        print(f"   Mat: {item['Matrícula']} | {item['Situação']} | {item['Percentual_Critico']}% crítico")
        print(f"   Obs Planilha: {item['Observação_Planilha'][:80] if item['Observação_Planilha'] else 'N/A'}")

if apenas_estudo:
    print("\n🔴 APENAS ESTUDO (Críticos não na planilha):")
    for i, item in enumerate(apenas_estudo[:5], 1):
        print(f"{i}. {item['Nome']} - Mat: {item['Matrícula']} - {item['Situação']} - {item['Percentual_Critico']}%")

if apenas_planilha:
    print("\n🔵 APENAS PLANILHA (Não críticos):")
    for i, item in enumerate(apenas_planilha[:5], 1):
        obs = item['Observação'][:50] if item['Observação'] else 'N/A'
        print(f"{i}. {item['Nome']} - Mat: {item['Matrícula']} - Obs: {obs}")

print("\n" + "="*80)
print("✅ ANÁLISE CONCLUÍDA!")
print("="*80)
